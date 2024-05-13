
import sys
from pathlib import Path
import re
import openpyxl

from pypdf import PdfReader

# TODO : compare nextcloud structure with handbook structure
# TODO : check broken links, missing OCR, non-incremental numbers, non-linked pages
# Update links/status in 30.02
# Create GitHub Issues (e.g., to update deadlines)

def compare_directories(nextcloud_path, docs_path):
    nextcloud_dirs = {p.name for p in nextcloud_path.iterdir() if p.is_dir()}
    docs_dirs = {p.name for p in docs_path.iterdir() if p.is_dir()}
    
    if nextcloud_dirs != docs_dirs:
        print("Warning: Directory structures of Nextcloud and docs are not identical.")
        print(f"Nextcloud directories: {nextcloud_dirs}")
        print(f"Docs directories: {docs_dirs}")
        sys.exit("Error")

def check_archive():

    archive_path = nextcloud_path / '10-lab' / '19_archive'
    pattern = re.compile(r'^\d{6}_\d{8}_[\w\s\-_]+\.pdf$')
    file_errors = False
    for file in archive_path.iterdir():
        if not pattern.match(file.name):
            file_errors = True
            print(f"Warning: File {file.name} does not match the expected pattern.")
    
        with open(file, 'rb') as f:
            reader = PdfReader(f)
            content = reader.pages[0].extract_text()
            if not content.strip():
                file_errors = True
                print(f"Nextcould warning: File {file.name} does not contain text.")

    # Check for missing incremental numbers in filenames
    file_numbers = sorted([int(file.name[:6]) for file in archive_path.iterdir() if pattern.match(file.name)])
    expected_numbers = list(range(min(file_numbers), max(file_numbers) + 1))
    expected_numbers.remove(23)  # 000023 is missing on purpose
    if file_numbers != expected_numbers:
        missing_numbers = set(expected_numbers) - set(file_numbers)
        file_errors = True
        print(f"Nextcould warning: Missing file(s) in 10-lab/19_archive with these incremental numbers: {missing_numbers}")

    if file_errors:
        sys.exit("Error")


def linkcheck():
    index_md_path = 'index.md'
    md_files = [index_md_path] + list(docs_path.glob('**/*.md'))
    link_pattern = re.compile(r'\[.*?\]\((.*?)\)')
    for md_file in md_files:
        with open(md_file, 'r', encoding='utf-8') as file:
            content = file.read()
            links = link_pattern.findall(content)
            for link in links:
                if link.startswith('http') or link.startswith("#") or link.startswith("mailto") or link.endswith("/"):
                    continue
                if link.startswith('{{ site.baseurl }}'):
                    link = link.replace('{{ site.baseurl }}', str(Path.cwd()))
                # print(f"Found link: {link}")
                # Replace .html with .md in the link
                link_md = link.split('#')[0].replace('.html', '.md')
                # Construct the full path for the link, removing everything right of "#"
                if str(md_file).startswith(' '):
                    print(f"Invalid link in handbook: In {md_file}: {link_md} starts with a space.")
                    continue
                if str(md_file).endswith(' '):
                    print(f"Invalid link in handbook: In {md_file}: {link_md} ends with a space.")
                    continue
                link_path = Path(md_file).parent / link_md
                # Check if the link is a relative path and if it exists
                if not link_path.is_absolute() and not link_path.exists():
                    print(f"Broken link in handbook: In {md_file}:  {link_md}")
                    

def check_theses():

    # Define the path to the excel file
    excel_path = nextcloud_path / '30-teaching' / '35_theses' / '000_overview' / '35.000 Theses.xlsx'
    # Load the workbook and select the sheets
    workbook = openpyxl.load_workbook(excel_path)
    bachelor_sheet = workbook['Bachelor-Arbeiten']
    master_sheet = workbook['Master-Arbeiten']

    # Create lists from the Bachelor-Arbeiten sheet
    bachelor_theses_names = []
    for row in bachelor_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0]:  # Ensure the cell is not empty
            bachelor_theses_names.append(row[0])

    # Create lists from the Master-Arbeiten sheet
    master_theses_names = []
    for row in master_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0]:  # Ensure the cell is not empty
            master_theses_names.append(row[0])

    # Define the path to the theses directory
    theses_dir = nextcloud_path / '30-teaching' / '35_theses'

    # Check if the directory exists
    if not theses_dir.exists():
        print(f"Directory {theses_dir} does not exist.")
        return

    # Get the list of all directories in the theses directory
    all_dirs = [d.name for d in theses_dir.iterdir() if d.is_dir()]

    # Check for each participant from bachelor/master theses
    for theses_names in [bachelor_theses_names, master_theses_names]:
        for name in theses_names:
            # Take the last name (before the comma)
            last_name = name.split(',')[0].strip()

            # Check whether there is a folder containing that name
            if not any(last_name in dir_name for dir_name in all_dirs):
                print(f"Theses in nextcloud/30-teaching/35_theses : No directory found for participant: {name}")


nextcloud_path = Path('nextcloud')
docs_path = Path('docs')

check_theses()
linkcheck()

if nextcloud_path.exists() and nextcloud_path.is_dir():
    compare_directories(nextcloud_path, docs_path)
    check_archive()
else:
    print("Nextcloud directory does not exist.")
    # TODO : error if not in github-actions
    sys.exit("Error")
